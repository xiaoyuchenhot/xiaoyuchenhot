import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ── Sheet 1: Incidents ──
ws = wb.active
ws.title = "Incidents"

headers = [
    "Case Number", "External Case Number", "Title", "Account", "Case Type",
    "Channel", "Priority", "Status", "Urgency", "Impact",
    "SLA Name", "SLA Time Remaining", "SLA Breached", "Days Open",
    "Activity Count", "Last Activity Date", "Escalation Count", "Email Count",
    "Phase", "Phase Change Count", "Assigned To", "Latest Comment",
    "Support Plan", "Preferred Language"
]

# Header styles
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
data_font = Font(name="Arial", size=10)
thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
red_font = Font(name="Arial", size=10, color="9C0006")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
yellow_font = Font(name="Arial", size=10, color="9C6500")
alt_fill_blue = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
alt_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# Sample data - 20 incidents with mixed risk levels
incidents = [
    # HIGH RISK cases (4)
    ["CS0741201", "CS0638979", "Unable to Core Share any documents", "Waikato Regional Council",
     "Technical Support Case", "Portal", "3 - Moderate", "Open", "2 - Medium", "2 - Medium",
     "ENT Follow-Up P3", "-3d 20h 3m", "Yes", 28, 147, "2026-07-28", 2, 10,
     "Root Cause Analysis", 5, "Bao Ha", "DTreeAncestors table has unexpected record for doc ID 35167451",
     "ENT Default Support", "English"],

    ["CS0741215", "CS0639102", "OTDS authentication failing for all SSO users - production down", "Deutsche Bank AG",
     "Technical Support Case", "Phone", "1 - Critical", "Open", "1 - High", "1 - High",
     "ENT Follow-Up P1", "-1d 8h 15m", "Yes", 5, 89, "2026-08-10", 3, 14,
     "Investigation", 4, "", "OTDS token validation returning 401 across all federated IdPs. Full SSO outage.",
     "ENT Premium Support", "English"],

    ["CS0741230", "CS0639245", "Archive Server retrieval timeout on large document sets", "Rio Tinto Mining",
     "Technical Support Case", "Email", "2 - High", "Pending Engineering", "1 - High", "2 - Medium",
     "ENT Follow-Up P2", "-5d 2h 40m", "Yes", 21, 63, "2026-07-30", 1, 8,
     "Fix Development", 6, "Evelyn Kawrykow", "Archive retrieval exceeds 120s timeout for collections >500 docs. Temp workaround: batch retrieval.",
     "ENT Premium Support", "English"],

    ["CS0741244", "CS0639301", "Content Server crash during scheduled workflow execution", "Telstra Corporation",
     "Technical Support Case", "Portal", "2 - High", "Open", "1 - High", "1 - High",
     "ENT Follow-Up P2", "-2d 14h 50m", "Yes", 12, 34, "2026-08-01", 2, 12,
     "Investigation", 3, "Marcus Chen", "CS crashes with OutOfMemoryError when workflow engine processes >200 concurrent assignments.",
     "ENT Premium Support", "English"],

    # MEDIUM RISK cases (5)
    ["CS0741256", "CS0639388", "Search indexing fails after Content Server patch 27.2.0.3340", "Ministry of Health NZ",
     "Technical Support Case", "Portal", "3 - Moderate", "Pending Support", "2 - Medium", "2 - Medium",
     "ENT Follow-Up P3", "1d 4h 20m", "No", 16, 42, "2026-08-07", 0, 6,
     "Testing", 4, "Sarah Mitchell", "Solr re-index completes but new documents not appearing in search results post-patch.",
     "ENT Default Support", "English"],

    ["CS0741263", "CS0639412", "WebReports dashboard rendering blank for specific user groups", "Deutsche Bank AG",
     "Technical Support Case", "Email", "3 - Moderate", "Open", "2 - Medium", "2 - Medium",
     "ENT Follow-Up P3", "0d 6h 45m", "No", 10, 28, "2026-08-09", 1, 7,
     "Root Cause Analysis", 3, "Bao Ha", "Permission inheritance issue on WebReports folder. Group ACL not propagating.",
     "ENT Premium Support", "English"],

    ["CS0741278", "CS0639450", "Records Management classification plan import error", "NSW Department of Education",
     "Technical Support Case", "Portal", "3 - Moderate", "Pending Engineering", "2 - Medium", "2 - Medium",
     "ENT Follow-Up P3", "2d 10h 30m", "No", 18, 55, "2026-08-04", 0, 4,
     "Fix Development", 5, "James Rodriguez", "XML import of classification plan fails at node depth >8. Schema validation passes.",
     "ENT Default Support", "English"],

    ["CS0741285", "CS0639478", "Email notification templates not rendering HTML correctly", "Auckland Transport",
     "Technical Support Case", "Email", "3 - Moderate", "Open", "2 - Medium", "3 - Low",
     "ENT Follow-Up P3", "3d 1h 15m", "No", 14, 19, "2026-08-02", 0, 6,
     "Investigation", 2, "Priya Sharma", "Notification emails showing raw HTML tags in Outlook. Works in webmail clients.",
     "ENT Default Support", "English"],

    ["CS0741292", "CS0639501", "Content Suite workspace template duplication creates orphan nodes", "Rio Tinto Mining",
     "Technical Support Case", "Portal", "3 - Moderate", "Pending Support", "2 - Medium", "2 - Medium",
     "ENT Follow-Up P3", "1d 18h 0m", "No", 9, 37, "2026-08-08", 1, 5,
     "Root Cause Analysis", 4, "Marcus Chen", "Duplicated workspace has 12 orphan nodes in DTreeCore. Parent references point to source workspace.",
     "ENT Premium Support", "English"],

    # LOW RISK cases (11)
    ["CS0741305", "CS0639534", "Need assistance configuring SAML 2.0 for OTDS", "Bank of Queensland",
     "Technical Support Case", "Portal", "4 - Low", "Pending Customer", "3 - Low", "3 - Low",
     "ENT Follow-Up P4", "12d 5h 30m", "No", 3, 8, "2026-08-11", 0, 2,
     "Investigation", 1, "Sarah Mitchell", "Sent SAML configuration guide and metadata template. Awaiting customer IdP metadata.",
     "ENT Default Support", "English"],

    ["CS0741312", "CS0639567", "Workflow map designer - save button greyed out intermittently", "Waikato Regional Council",
     "Technical Support Case", "Portal", "4 - Low", "Open", "3 - Low", "3 - Low",
     "ENT Follow-Up P4", "18d 12h 0m", "No", 5, 12, "2026-08-10", 0, 3,
     "Investigation", 1, "Priya Sharma", "Appears related to session timeout. Save button recovers after page refresh.",
     "ENT Default Support", "English"],

    ["CS0741325", "CS0639590", "Request to increase max upload file size from 100MB to 500MB", "Ministry of Health NZ",
     "Technical Support Case", "Email", "4 - Low", "Pending Customer", "3 - Low", "3 - Low",
     "ENT Follow-Up P4", "20d 0h 0m", "No", 7, 5, "2026-08-11", 0, 2,
     "Pending Support", 1, "James Rodriguez", "Provided configuration steps for IIS and Content Server upload limits. Awaiting change window.",
     "ENT Default Support", "English"],

    ["CS0741338", "CS0639623", "Transport Package export fails for packages >2GB", "NSW Department of Education",
     "Technical Support Case", "Portal", "3 - Moderate", "Pending Engineering", "2 - Medium", "2 - Medium",
     "ENT Follow-Up P3", "5d 8h 45m", "No", 8, 22, "2026-08-09", 0, 3,
     "Fix Development", 2, "Evelyn Kawrykow", "Known limitation in 27.1. Fix included in 27.2.1 patch. ETA 2 weeks.",
     "ENT Default Support", "English"],

    ["CS0741345", "CS0639656", "PDF rendition quality degraded after server migration", "Auckland Transport",
     "Technical Support Case", "Portal", "3 - Moderate", "Open", "2 - Medium", "3 - Low",
     "ENT Follow-Up P3", "8d 14h 20m", "No", 4, 15, "2026-08-10", 0, 2,
     "Investigation", 1, "Bao Ha", "Ghostscript version mismatch between old and new server. Recommending upgrade to 10.02.",
     "ENT Default Support", "English"],

    ["CS0741358", "CS0639689", "Connected Workspace sync delay between CS and Teams", "Telstra Corporation",
     "Technical Support Case", "Email", "3 - Moderate", "Pending Support", "2 - Medium", "2 - Medium",
     "ENT Follow-Up P3", "4d 22h 10m", "No", 6, 18, "2026-08-11", 0, 4,
     "Root Cause Analysis", 2, "Sarah Mitchell", "Sync queue backed up. Azure AD connector throttling detected. Adjusting batch size.",
     "ENT Premium Support", "English"],

    ["CS0741365", "CS0639712", "Content Server - slow query on Enterprise.DTreeNotify table", "Rio Tinto Mining",
     "Technical Support Case", "Portal", "3 - Moderate", "Pending Support", "2 - Medium", "2 - Medium",
     "ENT Follow-Up P3", "6d 10h 0m", "No", 7, 25, "2026-08-10", 0, 3,
     "Root Cause Analysis", 2, "Marcus Chen", "DTreeNotify table has 4.2M rows. Recommending archival of notifications older than 90 days.",
     "ENT Premium Support", "English"],

    ["CS0741378", "CS0639745", "OTDS user sync from Active Directory skipping nested groups", "Bank of Queensland",
     "Technical Support Case", "Portal", "3 - Moderate", "Open", "2 - Medium", "2 - Medium",
     "ENT Follow-Up P3", "5d 16h 30m", "No", 6, 14, "2026-08-11", 0, 3,
     "Investigation", 1, "Priya Sharma", "OTDS AD connector configured for single-level sync. Need to enable recursive group membership.",
     "ENT Default Support", "English"],

    ["CS0741385", "CS0639778", "Content Server - automated metadata assignment rule not firing", "Waikato Regional Council",
     "Technical Support Case", "Email", "3 - Moderate", "Pending Support", "2 - Medium", "3 - Low",
     "ENT Follow-Up P3", "7d 3h 45m", "No", 5, 16, "2026-08-10", 0, 2,
     "Root Cause Analysis", 2, "James Rodriguez", "Category inheritance rule has incorrect trigger condition. Rule fires on create but not on copy.",
     "ENT Default Support", "English"],

    ["CS0741392", "CS0639801", "SuccessFactors integration - employee termination not propagating", "Deutsche Bank AG",
     "Technical Support Case", "Portal", "2 - High", "Pending Support", "2 - Medium", "1 - High",
     "ENT Follow-Up P2", "3d 8h 20m", "No", 4, 20, "2026-08-11", 0, 4,
     "Investigation", 2, "Evelyn Kawrykow", "SAP SF webhook delivering events but OTDS user deprovisioning job runs daily, not real-time.",
     "ENT Premium Support", "English"],

    ["CS0741405", "CS0639834", "Content Server - version control merge conflict on co-authored docs", "Bank of Queensland",
     "Technical Support Case", "Email", "4 - Low", "Resolved", "3 - Low", "3 - Low",
     "ENT Follow-Up P4", "N/A", "No", 2, 6, "2026-08-12", 0, 1,
     "Done", 2, "Bao Ha", "Resolved by enabling optimistic locking and configuring auto-merge for minor edits.",
     "ENT Default Support", "English"],
]

for row_idx, incident in enumerate(incidents, 2):
    row_fill = alt_fill_blue if row_idx % 2 == 0 else alt_fill_white
    for col_idx, value in enumerate(incident, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [3, 22]))

    # Conditional formatting: SLA Breached column (13)
    sla_cell = ws.cell(row=row_idx, column=13)
    if sla_cell.value == "Yes":
        sla_cell.fill = red_fill
        sla_cell.font = red_font
        # Also highlight the SLA Time Remaining
        time_cell = ws.cell(row=row_idx, column=12)
        time_cell.fill = red_fill
        time_cell.font = red_font

    # Conditional formatting: Days Open > 14
    days_cell = ws.cell(row=row_idx, column=14)
    if isinstance(days_cell.value, int) and days_cell.value > 14:
        days_cell.fill = yellow_fill
        days_cell.font = yellow_font

# Column widths
col_widths = {
    1: 14, 2: 20, 3: 55, 4: 26, 5: 22, 6: 10, 7: 15, 8: 20,
    9: 13, 10: 13, 11: 18, 12: 20, 13: 14, 14: 11, 15: 14,
    16: 18, 17: 16, 18: 12, 19: 22, 20: 18, 21: 20, 22: 70,
    23: 22, 24: 18
}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# ── Sheet 2: Risk Scoring Guide ──
ws2 = wb.create_sheet("Risk Scoring Guide")

guide_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
guide_header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
section_font = Font(name="Arial", bold=True, size=12, color="1F4E79")
guide_font = Font(name="Arial", size=10)
guide_font_bold = Font(name="Arial", bold=True, size=10)

ws2.merge_cells("A1:D1")
title_cell = ws2.cell(row=1, column=1, value="Incident Risk Scoring Guide")
title_cell.font = Font(name="Arial", bold=True, size=16, color="1F4E79")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 35

ws2.merge_cells("A2:D2")
subtitle = ws2.cell(row=2, column=1, value="Used by Aviator Agent to classify incident risk levels")
subtitle.font = Font(name="Arial", italic=True, size=10, color="666666")
subtitle.alignment = Alignment(horizontal="center")

# Risk Signals table
row = 4
for col_idx, h in enumerate(["Signal", "Condition", "Points", "Rationale"], 1):
    cell = ws2.cell(row=row, column=col_idx, value=h)
    cell.font = guide_header_font
    cell.fill = guide_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

signals = [
    ["SLA Breached", "SLA Breached = Yes", "+3", "Direct contractual violation; customer trust at risk"],
    ["SLA Near Breach", "SLA < 20% time remaining", "+2", "Imminent breach if no action taken"],
    ["Overdue for Priority", "Days Open > threshold (P1: >3d, P2: >7d, P3: >14d, P4: >30d)", "+2", "Case exceeds expected resolution time for its severity"],
    ["Stale Case", "No activity in last 5 days", "+2", "Customer may feel abandoned; investigation has stalled"],
    ["High Email Volume", "Email count > 5 with status not Resolved", "+2", "Customer repeatedly reaching out; signal of frustration"],
    ["Escalated", "Escalation count > 0", "+1", "Case has been formally escalated at least once"],
    ["Phase Bouncing", "Phase change count > 3", "+1", "Case moving back and forth; no clear resolution path"],
]

for i, signal in enumerate(signals):
    r = row + 1 + i
    fill = alt_fill_blue if i % 2 == 0 else alt_fill_white
    for col_idx, val in enumerate(signal, 1):
        cell = ws2.cell(row=r, column=col_idx, value=val)
        cell.font = guide_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    # Bold the points column
    ws2.cell(row=r, column=3).font = guide_font_bold

# Classification table
row = 14
ws2.cell(row=row, column=1, value="Risk Classification").font = section_font

row = 15
for col_idx, h in enumerate(["Risk Level", "Score Range", "Action Required"], 1):
    cell = ws2.cell(row=row, column=col_idx, value=h)
    cell.font = guide_header_font
    cell.fill = guide_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

classifications = [
    ["HIGH RISK", "5+ points", "Immediate manager review. Contact customer within 2 hours. Assign senior engineer."],
    ["MEDIUM RISK", "3-4 points", "Review within 24 hours. Ensure active ownership and next steps documented."],
    ["LOW RISK", "0-2 points", "Normal workflow. Monitor in next scheduled review cycle."],
]

risk_fills = [
    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
]
risk_fonts = [
    Font(name="Arial", bold=True, size=10, color="9C0006"),
    Font(name="Arial", bold=True, size=10, color="9C6500"),
    Font(name="Arial", bold=True, size=10, color="006100"),
]

for i, cls in enumerate(classifications):
    r = row + 1 + i
    for col_idx, val in enumerate(cls, 1):
        cell = ws2.cell(row=r, column=col_idx, value=val)
        cell.font = risk_fonts[i] if col_idx == 1 else guide_font
        cell.fill = risk_fills[i]
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# Priority thresholds
row = 21
ws2.cell(row=row, column=1, value="Priority-Based Thresholds").font = section_font

row = 22
for col_idx, h in enumerate(["Priority", "Max Days Open", "SLA Target"], 1):
    cell = ws2.cell(row=row, column=col_idx, value=h)
    cell.font = guide_header_font
    cell.fill = guide_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

thresholds = [
    ["1 - Critical", "3 days", "4 hours response, 24 hours resolution"],
    ["2 - High", "7 days", "8 hours response, 3 days resolution"],
    ["3 - Moderate", "14 days", "24 hours response, 7 days resolution"],
    ["4 - Low", "30 days", "48 hours response, 30 days resolution"],
]

for i, t in enumerate(thresholds):
    r = row + 1 + i
    fill = alt_fill_blue if i % 2 == 0 else alt_fill_white
    for col_idx, val in enumerate(t, 1):
        cell = ws2.cell(row=r, column=col_idx, value=val)
        cell.font = guide_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

# Column widths for sheet 2
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 50
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 65

output_path = "/home/user/xiaoyuchenhot/incident_risk_monitor_sample_data.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
